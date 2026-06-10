import csv
import os
from datetime import datetime, date

import openpyxl

from app.database import get_db

EXCEL_PATH = '/app/Pasitos_Registro_Cursos.xlsx'
FORMATOS_SOPORTADOS = {'.xlsx', '.xls', '.csv'}
HOJAS_REQUERIDAS    = ['Catálogo de Cursos', 'Registro de Inscripciones']

# Columnas clave para detectar tipo de hoja / CSV
_COLS_CURSOS   = {'ID Curso', 'Nombre del Curso'}
_COLS_PERSONAS = {'CURP', 'Nombre Completo'}


# ── helpers generales ─────────────────────────────────────────────────────────

def _str(val, default=''):
    if val is None:
        return default
    if isinstance(val, (datetime, date)):
        return val.strftime('%d/%m/%Y')
    return str(val).strip()


def _build_col_map(worksheet, header_row):
    col_map = {}
    for i, cell in enumerate(worksheet[header_row]):
        if cell.value is not None:
            col_map[str(cell.value).strip()] = i
    return col_map


def _get(row, col_map, col_name, default=''):
    idx = col_map.get(col_name)
    if idx is None or idx >= len(row):
        return default
    return _str(row[idx], default)


def _encontrar_fila_encabezado(ws, columnas_clave, max_filas=15):
    """Devuelve el número de fila (1-based) donde aparecen los encabezados, o None."""
    for row_num in range(1, max_filas + 1):
        try:
            row_vals = {str(c.value or '').strip() for c in ws[row_num]}
            if columnas_clave & row_vals:
                return row_num
        except Exception:
            continue
    return None


# ── cargadores por formato ────────────────────────────────────────────────────

def _abrir_xlsx(filepath):
    return openpyxl.load_workbook(filepath, data_only=True)


def _abrir_xls(filepath):
    """Convierte .xls → openpyxl Workbook en memoria usando xlrd."""
    try:
        import xlrd
    except ImportError:
        raise RuntimeError(
            'La librería xlrd no está instalada. '
            'Agrega xlrd>=2.0.1 a requirements.txt y reconstruye la imagen.'
        )
    xls = xlrd.open_workbook(filepath)
    wb  = openpyxl.Workbook()
    wb.remove(wb.active)                        # quitar hoja vacía por defecto
    for sheet_name in xls.sheet_names():
        ws_xls = xls.sheet_by_name(sheet_name)
        ws     = wb.create_sheet(title=sheet_name)
        for r in range(ws_xls.nrows):
            for c in range(ws_xls.ncols):
                ws.cell(row=r + 1, column=c + 1, value=ws_xls.cell_value(r, c))
    return wb


def _leer_csv(filepath):
    """Lee un CSV probando distintos encodings. Retorna lista de listas."""
    for encoding in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
        try:
            with open(filepath, newline='', encoding=encoding) as f:
                return list(csv.reader(f))
        except UnicodeDecodeError:
            continue
    raise ValueError('No se pudo decodificar el CSV. Guárdalo como UTF-8 e inténtalo de nuevo.')


def _detectar_hoja_csv(rows):
    """Determina si el CSV es de cursos o inscripciones según los encabezados."""
    for row in rows[:15]:
        row_set = {str(v).strip() for v in row}
        if _COLS_CURSOS & row_set:
            return 'Catálogo de Cursos'
        if _COLS_PERSONAS & row_set:
            return 'Registro de Inscripciones'
    return None


def _abrir_csv(filepath):
    """Detecta el tipo de CSV y lo convierte a openpyxl Workbook en memoria."""
    rows       = _leer_csv(filepath)
    sheet_name = _detectar_hoja_csv(rows)
    if not sheet_name:
        raise ValueError(
            'No se encontraron encabezados válidos en el CSV. '
            'Debe contener "ID Curso"/"Nombre del Curso" (catálogo) '
            'o "CURP"/"Nombre Completo" (inscripciones).'
        )
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title=sheet_name)
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val or None)
    return wb


def _cargar_workbook(filepath):
    """Abre cualquier formato soportado y retorna un openpyxl Workbook."""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.xlsx':
        return _abrir_xlsx(filepath)
    if ext == '.xls':
        return _abrir_xls(filepath)
    if ext == '.csv':
        return _abrir_csv(filepath)
    raise ValueError(f'Formato no soportado: {ext}')


# ── validación ────────────────────────────────────────────────────────────────

def validar_formato(filepath):
    """Verifica que el archivo sea válido para importar.
    Retorna lista de errores (vacía si todo está correcto)."""
    errores = []
    ext = os.path.splitext(filepath)[1].lower()

    if ext not in FORMATOS_SOPORTADOS:
        errores.append(
            f'Formato no soportado. Se aceptan: {", ".join(sorted(FORMATOS_SOPORTADOS))}'
        )
        return errores

    if ext == '.csv':
        try:
            rows = _leer_csv(filepath)
            if not _detectar_hoja_csv(rows):
                errores.append(
                    'No se encontraron encabezados válidos. '
                    'El CSV debe tener "ID Curso"/"Nombre del Curso" '
                    'o "CURP"/"Nombre Completo".'
                )
        except Exception as e:
            errores.append(str(e))
        return errores

    # .xlsx / .xls → verificar que existan ambas hojas requeridas
    try:
        wb = _cargar_workbook(filepath)
        for hoja in HOJAS_REQUERIDAS:
            if hoja not in wb.sheetnames:
                errores.append(f'Falta la hoja "{hoja}"')
    except Exception as e:
        errores.append(f'No se pudo leer el archivo: {e}')
    return errores


# ── importación ───────────────────────────────────────────────────────────────

def importar_desde_excel(filepath=EXCEL_PATH):
    resultado = {'personas': 0, 'cursos': 0, 'errores': []}

    if not os.path.exists(filepath):
        resultado['errores'].append(f'Archivo no encontrado: {filepath}')
        return resultado

    try:
        wb = _cargar_workbook(filepath)
    except Exception as e:
        resultado['errores'].append(f'Error abriendo archivo: {e}')
        return resultado

    conn = get_db()

    # --- Catálogo de Cursos ---
    sheet_cursos = 'Catálogo de Cursos'
    if sheet_cursos in wb.sheetnames:
        ws         = wb[sheet_cursos]
        header_row = _encontrar_fila_encabezado(ws, _COLS_CURSOS) or 3
        col_map    = _build_col_map(ws, header_row)
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or row[0] is None:
                continue
            try:
                id_curso = _get(row, col_map, 'ID Curso')
                nombre   = _get(row, col_map, 'Nombre del Curso')
                if not id_curso or not nombre:
                    continue
                tipo         = _get(row, col_map, 'Tipo / Formato')
                duracion_raw = row[col_map['Duración (horas)']] if 'Duración (horas)' in col_map else None
                duracion     = int(float(duracion_raw)) if duracion_raw is not None else None
                modalidad    = _get(row, col_map, 'Modalidad')
                descripcion  = _get(row, col_map, 'Descripción')
                estado       = _get(row, col_map, 'Estado') or 'Activo'

                conn.execute(
                    '''INSERT OR IGNORE INTO cursos
                       (id, nombre, tipo, duracion_horas, modalidad, descripcion, estado)
                       VALUES (?, ?, ?, ?, ?, ?, ?)''',
                    (id_curso, nombre, tipo, duracion, modalidad, descripcion, estado),
                )
                resultado['cursos'] += 1
            except Exception as e:
                resultado['errores'].append(f'Curso: {e}')
    else:
        resultado['errores'].append(f'Hoja "{sheet_cursos}" no encontrada')

    # --- Registro de Inscripciones ---
    sheet_reg = 'Registro de Inscripciones'
    if sheet_reg in wb.sheetnames:
        ws         = wb[sheet_reg]
        header_row = _encontrar_fila_encabezado(ws, _COLS_PERSONAS) or 4
        col_map    = _build_col_map(ws, header_row)
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or row[0] is None:
                continue
            try:
                nombre = _get(row, col_map, 'Nombre Completo')
                curp   = _get(row, col_map, 'CURP').upper()
                if not nombre or not curp:
                    continue
                if len(curp) != 18:
                    resultado['errores'].append(f'CURP inválida ({len(curp)} chars): {curp}')
                    continue
                fecha_nac  = _get(row, col_map, 'Fecha de Nacimiento')
                grado      = _get(row, col_map, 'Último Grado de Estudio')
                correo     = _get(row, col_map, 'Correo Electrónico')
                institucion = _get(row, col_map, 'Institución / Guardería')
                cargo      = _get(row, col_map, 'Cargo o Puesto')

                conn.execute(
                    '''INSERT OR IGNORE INTO personas
                       (nombre, curp, fecha_nacimiento, grado_estudio, correo, institucion, cargo)
                       VALUES (?, ?, ?, ?, ?, ?, ?)''',
                    (nombre, curp, fecha_nac, grado, correo, institucion, cargo),
                )
                resultado['personas'] += 1
            except Exception as e:
                resultado['errores'].append(f'Persona: {e}')
    else:
        resultado['errores'].append(f'Hoja "{sheet_reg}" no encontrada')

    conn.commit()
    conn.close()
    return resultado
